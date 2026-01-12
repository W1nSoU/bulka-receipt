from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.db import Receipt, User


HEADER = [
    "ID",
    "ПІБ",
    "Телефон",
    "Telegram",
    "Магазин",
    "Сума",
    "Дата",
    "Час",
    "Код чеку",
    "Фото file_id",
]


def _get_sheet(path: Path) -> Worksheet:
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Checks"
        ws.append(HEADER)
        wb.save(path)
        return ws
    wb = load_workbook(path)
    ws = wb["Checks"] if "Checks" in wb.sheetnames else wb.active
    return ws


def ensure_workbook(path: Path) -> None:
    _get_sheet(path).parent.save(path)


def append_receipt(path: Path, receipt: Receipt, user: User, username: str | None) -> None:
    wb = load_workbook(path) if path.exists() else Workbook()
    ws = wb["Checks"] if "Checks" in wb.sheetnames else wb.active
    if ws.max_row == 1 and ws[1][0].value != HEADER[0]:
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER)
    ws.append(
        [
            receipt.id,
            user.full_name,
            user.phone,
            f"@{username}" if username else str(user.telegram_id),
            receipt.shop,
            receipt.amount,
            receipt.date,
            receipt.time,
            receipt.check_code,
            receipt.file_id,
        ]
    )
    wb.save(path)
